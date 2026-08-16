"""
Session workbook builder — one .xlsx per finished session, readable in
LibreOffice Calc.

Called two ways:
  - by the watcher, once a session has been idle long enough (see sessions.py)
  - as a CLI, to regenerate one session or backfill the whole history:

      python exporter.py --last
      python exporter.py --session 2026-08-10_2015_Pseudo
      python exporter.py --rebuild-all --out /volume1/.../exports

xlsx rather than ods: LibreOffice reads it natively, and openpyxl can write
charts, number formats and conditional formatting that the ods writers cannot.

Only the bankroll block holds live formulas. Everything else is written as a
value, so the numbers are right the moment the file opens, whatever the
spreadsheet's recalculation-on-load setting happens to be. The bankroll cells
are the exception on purpose: they hang off one figure, and typing into it
recalculates them.

That figure no longer has to be typed. When a bankroll ledger is configured
(BANKROLL_FILE, or --bankroll), the session is also written into it — one row
per session, see bankroll.py — and the workbook opens with the bankroll this
session actually started from.

The sheets are ordered by how much a single session can actually support:

  Résumé      the money, immediately followed by what that money is worth
  Formats     vanilla / knockout / mystery kept apart, never added up
  Comparatif  session rates against the player's own history, with sample sizes
  Profondeur  the same rates split by stack depth — three different games
  Sorties     how each tournament ended, deepest exit first
  Tournois    the raw table, one row per tournament
  Positions, Mains marquantes, Graphiques, Notes
"""

import argparse
import os
import re
import sys
from datetime import datetime, timezone
from math import sqrt
from statistics import median
from zoneinfo import ZoneInfo

import psycopg2
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import bankroll
import sessions

# ── Config ────────────────────────────────────────────────────────────────────

TZ = ZoneInfo("Europe/Paris")   # timestamps are stored as naive UTC

MONEY = '#,##0.00\\ "€"'
CHIPS = '#,##0'
COUNT = '#,##0'
PCT = '0.0%'
PCT_SIGNED = '+0.0%;-0.0%;0.0%'
RATIO = '0.00'
RATIO_SIGNED = '+0.00;-0.00;0.00'
BB = '+#,##0.0;-#,##0.0;0.0'
BB_UNIT = '#,##0.0" bb"'
BUYINS = '+0.0" caves";-0.0" caves";0.0" caves"'
SIGMA = '+0.00" σ";-0.00" σ";0.00" σ"'
DATETIME = 'DD/MM/YYYY HH:MM'
TIME = 'HH:MM'

# How a format is named on the sheets, and the caveat that goes with it.
FORMAT_LABELS = {
    "vanilla": "Vanilla",
    "knockout": "Knockout",
    "mystery": "Mystery",
    "sng": "Expresso / SNG",
    "inconnu": "Résumé absent",
}
FORMAT_NOTES = {
    "vanilla": "prize pool seul",
    "knockout": "une part du buy-in revient en primes, encaissées tout du long",
    "mystery": "primes tirées à partir d'un certain stade : gains très concentrés",
    "sng": "3 joueurs, hyper-turbo : compté dans l'argent, exclu de tout le reste",
    "inconnu": "buy-in et gains pas encore connus",
}
BOUNTY_FAMILIES = ("knockout", "mystery")

# Denominator below which a rate is a coin flip rather than a read.
SAMPLE_SOLID = 100
SAMPLE_INDICATIVE = 30

# ── Styling ───────────────────────────────────────────────────────────────────

TITLE_FONT = Font(bold=True, size=15)
SUBTITLE_FONT = Font(size=11, color="595959")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="2F4F5F")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4E7A8C")
LABEL_FONT = Font(size=10)
VALUE_FONT = Font(bold=True, size=10)
INPUT_FONT = Font(bold=True, size=11, color="7F6000")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FONT = Font(bold=True, size=10, color="9C0006")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
TOTAL_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EDF3F5")
NOTE_FONT = Font(italic=True, size=9, color="808080")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TAB_KEY = "2F4F5F"      # the sheets a session can actually support
TAB_DETAIL = "4E7A8C"   # the raw tables behind them
TAB_FREE = "BFBFBF"     # notes and charts


def _profit_scale() -> ColorScaleRule:
    """Red below zero, green above. A fresh rule per sheet — openpyxl numbers
    the rules it is given, and a shared instance ends up on two sheets at once."""
    return ColorScaleRule(
        start_type="min", start_color="F4A9A8",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="A9D08E",
    )


# ── Small helpers ─────────────────────────────────────────────────────────────

def _f(value) -> float | None:
    """Decimal → float. NUMERIC columns come back as Decimal, which mixes badly
    with plain arithmetic further down."""
    return None if value is None else float(value)


def _div(numerator, denominator) -> float | None:
    """Rate or None — the shape every ratio on these sheets needs."""
    numerator, denominator = _f(numerator), _f(denominator)
    if not denominator or numerator is None:
        return None
    return numerator / denominator


def _local(ts: datetime | None) -> datetime | None:
    """Naive UTC → naive Paris, for display."""
    if ts is None:
        return None
    return ts.replace(tzinfo=timezone.utc).astimezone(TZ).replace(tzinfo=None)


def _duration(seconds) -> str:
    if not seconds:
        return "—"
    seconds = int(seconds)
    return f"{seconds // 3600} h {seconds % 3600 // 60:02d}"


def _cards(raw: str | None) -> str:
    """'AhKc' → 'Ah Kc'."""
    if not raw:
        return ""
    return " ".join(raw[i:i + 2] for i in range(0, len(raw), 2))


def _safe(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9À-ÿ+._-]", "_", name)


def _autosize(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _header_row(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 28


def _data_row(ws, row: int, values: list, formats: list[str | None]) -> None:
    for col, (value, fmt) in enumerate(zip(values, formats), start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BOX
        if fmt:
            cell.number_format = fmt


def _banner(ws, row: int, text: str, width: int) -> None:
    """A section strip spanning `width` columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def _note(ws, row: int, text: str, column: int = 1) -> None:
    ws.cell(row=row, column=column, value=text).font = NOTE_FONT


def _printable(ws) -> None:
    """One page wide — these sheets are meant to be read, sometimes on paper."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _plural(count, singular: str, plural: str | None = None) -> str:
    return singular if abs(count) < 2 else (plural or singular + "s")


def _reliability(sample) -> str:
    """What a denominator is worth, said in words rather than implied."""
    if not sample:
        return "aucune occasion"
    if sample >= SAMPLE_SOLID:
        return f"n = {sample:,.0f} — exploitable".replace(",", " ")
    if sample >= SAMPLE_INDICATIVE:
        return f"n = {sample} — indicatif"
    return f"n = {sample} — bruit"


# ── Résumé ────────────────────────────────────────────────────────────────────

class _KeyValueSheet:
    """Label in column A, value in column B, sections in between."""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def title(self, text: str, subtitle: str = "") -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
        cell = self.ws.cell(row=self.row, column=1, value=text)
        cell.font = TITLE_FONT
        self.row += 1
        if subtitle:
            self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
            cell = self.ws.cell(row=self.row, column=1, value=subtitle)
            cell.font = SUBTITLE_FONT
            self.row += 1
        self.row += 1

    def section(self, text: str) -> None:
        _banner(self.ws, self.row, text, 3)
        self.row += 1

    def kv(self, label: str, value, fmt: str | None = None, note: str = "", style: str = "") -> str:
        """Writes one label/value pair and returns the value's cell reference."""
        self.ws.cell(row=self.row, column=1, value=label).font = LABEL_FONT
        cell = self.ws.cell(row=self.row, column=2, value=value)
        cell.font = VALUE_FONT
        cell.border = BOX
        if fmt:
            cell.number_format = fmt
        if style == "input":
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
        elif style == "warn":
            cell.font = WARN_FONT
            cell.fill = WARN_FILL
        if note:
            self.ws.cell(row=self.row, column=3, value=note).font = NOTE_FONT
        ref = f"$B${self.row}"
        self.row += 1
        return ref

    def text(self, content: str) -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
        self.ws.cell(row=self.row, column=1, value=content).font = NOTE_FONT
        self.row += 1

    def blank(self, n: int = 1) -> None:
        self.row += n


def _exit_depths(exits: list[dict], family: str | None = None) -> list[float]:
    """
    Stack in big blinds carried into the last hand, tournaments won excluded —
    winning one is not a way of going out, and it would drag the median down.

    With no family given the answer covers multi-table tournaments only: an
    Expresso ends at a depth its structure decided, not the player.
    """
    wanted = (family,) if family else sessions.MTT_FAMILIES
    return [
        _f(e["exit_bb"]) for e in exits
        if e["exit_bb"] is not None
        and not e["won_tournament"]
        and e["family"] in wanted
    ]


def _sheet_resume(wb: Workbook, data: dict) -> None:
    meta = data["meta"]
    start, end = _local(meta["session_start"]), _local(meta["session_end"])
    elapsed = (meta["session_end"] - meta["session_start"]).total_seconds()
    hours = elapsed / 3600 or None

    invested = _f(meta["total_invested"]) or 0.0
    won = _f(meta["total_won"]) or 0.0
    profit = _f(meta["net_profit"]) or 0.0
    tournaments = meta["tournaments_count"]
    hands = meta["hands_count"]
    paid = meta["paid_count"] or 0
    avg_buyin = _f(meta["avg_buyin"]) or 0.0
    profit_buyins = profit / avg_buyin if avg_buyin else None
    free = tournaments - paid

    formats = data["formats"]
    played = " + ".join(FORMAT_LABELS.get(f["family"], f["family"])
                        for f in formats if f["family"] != "inconnu")

    s = _KeyValueSheet(wb.create_sheet("Résumé"))
    s.ws.sheet_properties.tabColor = TAB_KEY
    _printable(s.ws)
    _autosize(s.ws, {"A": 36, "B": 18, "C": 58})

    s.title(
        f"Session n°{data['index']} — {start:%d/%m/%Y}",
        f"{meta['player']} · {start:%H:%M} → {end:%H:%M} · {_duration(elapsed)} · "
        f"{tournaments} tournois · {played}",
    )

    # Bankroll first: it is the only block that expects to be typed into, so it
    # sits where the file opens rather than buried under the results. The
    # figure comes pre-filled when the ledger could be read — it knows what the
    # bankroll was worth the moment this session started, which is exactly what
    # this cell used to ask the player to remember.
    s.section("BANKROLL")
    opening = data.get("bankroll_start")
    bankroll_ref = s.kv(
        "Bankroll de départ", opening, MONEY,
        "lue dans le fichier bankroll" if opening is not None else "← à saisir",
        style="input",
    )
    caves = s.kv("Règle de gestion (nb de caves)", 100, COUNT, "← à ajuster", style="input")
    s.blank()

    s.section("RÉSULTAT")
    s.kv("Tournois joués", tournaments, COUNT,
         f"dont {meta['reentries']} {_plural(meta['reentries'], 're-entry', 're-entries')}"
         if meta["reentries"] else "")
    buyin_ref = s.kv(
        "Buy-in moyen", avg_buyin, MONEY,
        f"sur les {paid} entrées payantes — {free} gratuite"
        f"{_plural(free, '', 's')} (freeroll, ticket, day 2)" if free else "",
    )
    s.kv("Total investi", invested, MONEY)
    s.kv("Total encaissé", won, MONEY)
    profit_ref = s.kv("Profit net", profit, MONEY)
    # The unit that survives a session mixing stakes, and the one to compare
    # with any other session.
    s.kv("Profit en caves", profit_buyins, BUYINS, "profit rapporté au buy-in moyen")
    s.kv("ROI de la session", profit / invested if invested else None, PCT_SIGNED)
    s.kv("Profit cumulé (depuis le début)", data["lifetime_profit"], MONEY,
         "toutes sessions confondues")
    s.blank()

    _resume_variance(s, data, profit_buyins, avg_buyin)
    _resume_runs(s, data)
    _resume_formats(s, data)

    # Placed after the profit cell it references — a spreadsheet does not care
    # about ordering, and the bankroll belongs at the top for the eye.
    s.section("GESTION DE BANKROLL")
    end_ref = s.kv(
        "Bankroll de fin",
        f'=IFERROR(IF({bankroll_ref}="","—",{bankroll_ref}+{profit_ref}),"—")',
        MONEY,
    )
    s.kv(
        "Variation",
        f'=IFERROR(IF({bankroll_ref}="","—",{profit_ref}/{bankroll_ref}),"—")',
        PCT_SIGNED,
    )
    s.kv(
        "Part de bankroll engagée",
        f'=IFERROR(IF({bankroll_ref}="","—",{invested:.2f}/{bankroll_ref}),"—")',
        PCT,
    )
    s.kv(
        "Caves couvertes en fin de session",
        f'=IFERROR(IF({bankroll_ref}="","—",{end_ref}/{buyin_ref}),"—")',
        COUNT,
        "à ce buy-in moyen",
    )
    s.kv(
        "Buy-in max selon la règle",
        f'=IFERROR(IF(OR({bankroll_ref}="",{caves}=0),"—",{end_ref}/{caves}),"—")',
        MONEY,
    )
    s.blank()

    s.section("VOLUME")
    s.kv("Durée de la session", _duration(elapsed), None, "première à dernière main")
    s.kv(
        "Temps de table cumulé",
        _duration(meta["played_seconds"]),
        None,
        "somme des tournois — dépasse la durée si multitable",
    )
    s.kv("Mains jouées", hands, CHIPS,
         f"dont {meta['mtt_hands_count']} en MTT" if meta["mtt_hands_count"] != hands else "")
    s.kv("Mains par heure", hands / hours if hours else None, COUNT)
    s.kv("Tournois par heure", tournaments / hours if hours else None, RATIO)

    missing = meta["tournaments_count"] - meta["summaries_count"]
    if missing:
        s.blank()
        s.section("ATTENTION")
        s.kv(
            "Résumés manquants",
            missing,
            None,
            "buy-in et gains incomplets — le classeur sera régénéré à leur arrivée",
            style="warn",
        )


def _resume_variance(s: _KeyValueSheet, data: dict, profit_buyins, avg_buyin) -> None:
    """
    What the session's money is worth as evidence.

    A night of MTTs is a handful of draws from a distribution whose whole mass
    sits in a tail nobody hits twice a week. Expressing the result in standard
    deviations is the only honest way to print it: it says, in one number,
    whether anything happened at all.
    """
    var = data["variance"]
    n = data["meta"]["tournaments_count"] or 0
    sd = var["sd_buyins"]                      # per tournament, in buy-ins
    noise = sd * sqrt(n) if n else None        # for the session as a whole
    z = profit_buyins / noise if noise and profit_buyins is not None else None

    if z is None:
        verdict, style = "—", ""
    elif abs(z) < 1:
        verdict, style = "dans le bruit", ""
    elif abs(z) < 2:
        verdict, style = "écart ordinaire", ""
    else:
        verdict, style = "gros écart", "warn"

    s.section("CE QUE CE RÉSULTAT VAUT")
    s.kv(
        "Écart-type par tournoi",
        sd,
        '0.00" caves"',
        "valeur par défaut, historique trop court" if var["borrowed"]
        else f"mesuré sur tes {var['sample']} tournois",
    )
    s.kv("Bruit attendu sur la session", noise, '0.0" caves"', "à 1 σ, dans un sens ou dans l'autre")
    s.kv("soit en euros", noise * avg_buyin if noise and avg_buyin else None, MONEY, "± à 1 σ")
    s.kv("Position du résultat", z, SIGMA, verdict, style=style)
    s.kv(
        "Tournois pour un ROI à ±10 points",
        (sd / 0.10) ** 2 if sd else None,
        COUNT,
        "à ton écart-type — c'est l'ordre de grandeur d'une saison",
    )
    s.text(
        "Un résultat sous 1 σ ne dit rien de la façon dont tu as joué, ni dans un sens "
        "ni dans l'autre : c'est l'amplitude normale du hasard sur ce nombre de tournois. "
        "Les onglets suivants sont là pour ça."
    )
    s.blank()


def _resume_runs(s: _KeyValueSheet, data: dict) -> None:
    """
    The part of a session that is nearly free of variance.

    Where you finish in the field is known for every tournament, cashed or not,
    and needs no prize to exist — which makes it the lowest-noise thing a
    single session produces. The exit depth is the other one: it is a fact
    about the last decision, not about how it turned out.
    """
    meta = data["meta"]
    n = meta["mtt_count"] or 0
    if not n:
        return

    depths = _exit_depths(data["exits"])
    deep_exits = [d for d in depths if d >= sessions.DEEP_EXIT_BB]
    others = (meta["tournaments_count"] or 0) - n

    s.section("QUALITÉ DES PARCOURS" + (f" — {n} MTT sur {n + others}" if others else ""))
    if others:
        s.text(f"{others} {_plural(others, 'tournoi')} hors MTT — Expresso et assimilés — "
               f"{_plural(others, 'est', 'sont')} compté{_plural(others, '', 's')} dans "
               "l'argent plus haut, et exclu"
               f"{_plural(others, '', 's')} de tout ce qui suit.")
    s.kv("Place médiane", _f(meta["median_finish_pct"]), PCT,
         "en % du champ — plus c'est bas, mieux c'est")
    s.kv("Meilleur parcours", _f(meta["best_finish_pct"]), PCT, "meilleure place de la session")
    s.kv("ITM", _div(meta["itm"], n), PCT,
         f"{meta['itm']} sur {n} — le prize pool a payé")
    if meta["cashed"] != meta["itm"]:
        s.kv("Tournois avec du gain", _div(meta["cashed"], n), PCT,
             f"{meta['cashed']} sur {n} — primes comprises")
    s.kv(f"Top {sessions.DEEP_RUN_PCT:.0%} du champ".replace("%", " %"),
         _div(meta["deep_runs"], n), PCT,
         f"{meta['deep_runs']} sur {n} — c'est là que se fait le résultat")
    s.kv(f"Tables finales (top {sessions.FINAL_TABLE_PLACES})", meta["final_tables"], COUNT)
    s.kv("Sortie médiane", median(depths) if depths else None, BB_UNIT,
         "tapis emporté dans la dernière main")
    s.kv(
        f"Sorties au-dessus de {sessions.DEEP_EXIT_BB} bb",
        len(deep_exits),
        COUNT,
        "à revoir en priorité : un tapis choisi, pas subi" if deep_exits
        else "aucune — les sorties sont venues à tapis court",
        style="warn" if deep_exits else "",
    )
    s.blank()


def _resume_formats(s: _KeyValueSheet, data: dict) -> None:
    """A line per format, with the detail one sheet away."""
    if len(data["formats"]) < 2:
        return

    s.section("PAR FORMAT")
    for raw in data["formats"]:
        view = _format_view(raw, data["exits"])
        label = FORMAT_LABELS.get(raw["family"], raw["family"])
        roi = view["roi"]
        s.kv(
            f"{label} — {view['tournaments']} {_plural(view['tournaments'], 'tournoi')}",
            view["profit"],
            MONEY,
            f"ROI {roi:+.0%} · ITM {view['itm_n']}/{view['tournaments']}"
            if roi is not None else "buy-in inconnu",
        )
    s.text("Détail, primes comprises, dans l'onglet Formats.")
    s.blank()


# ── Formats ───────────────────────────────────────────────────────────────────

# (label, view key, number format, note). "§" opens a section.
_FORMAT_ROWS = [
    ("§", "VOLUME", None, ""),
    ("Tournois joués", "tournaments", COUNT, ""),
    ("Re-entries payées", "reentries", COUNT, "entrées supplémentaires dans un même tournoi"),
    ("Buy-in moyen", "avg_buyin", MONEY, ""),
    ("Champ moyen", "avg_field", COUNT, "joueurs inscrits"),
    ("Mains jouées", "hands", COUNT, ""),
    ("Temps de table", "played", None, "somme des tournois"),
    ("§", "ARGENT", None, ""),
    ("Investi", "invested", MONEY, ""),
    ("Encaissé", "won", MONEY, ""),
    ("Profit", "profit", MONEY, ""),
    ("Profit en caves", "profit_buyins", BUYINS, "rapporté au buy-in moyen du format"),
    ("ROI", "roi", PCT_SIGNED, "sur si peu de tournois, à lire comme une anecdote"),
    ("Rake payé", "rake", MONEY, "part du buy-in qui ne joue pas"),
    ("§", "PARCOURS", None, ""),
    ("ITM", "itm_pct", PCT, "le prize pool a payé — primes exclues"),
    ("Top 10 % du champ", "deep_pct", PCT, "la seule zone qui finance un ROI en MTT"),
    ("Tables finales", "final_tables", COUNT, "top 9"),
    ("Place médiane", "median_finish", PCT, "en % du champ"),
    ("Sortie médiane", "median_exit", BB_UNIT, "tapis emporté dans la dernière main"),
    ("§", "PRIMES", None, ""),
    ("Part primes du buy-in", "bounty_share_buyin", PCT, "ce que tu paies pour la partie primes"),
    ("Primes encaissées", "bounty_won", MONEY, ""),
    ("Solde primes", "bounty_balance", MONEY, "primes encaissées – primes payées"),
    ("Solde prize pool", "prize_balance", MONEY, "gains prize – part prize du buy-in"),
    ("Part des gains en primes", "bounty_share_won", PCT, ""),
    ("§", "JEU", None, ""),
    ("VPIP", "vpip", PCT, ""),
    ("PFR", "pfr", PCT, ""),
    ("Flop vu", "flop", PCT, ""),
    ("3-bet", "three_bet", PCT, ""),
]


def _format_view(raw: dict, exits: list[dict]) -> dict:
    """
    Turn one format's counters into the figures the sheet prints.

    Rates are rebuilt from numerators and denominators here rather than in SQL,
    which is what lets the "toutes formes" column be the sum of the counters
    instead of an average of percentages.
    """
    n = raw["tournaments"] or 0
    # A family whose summaries have not landed has no money at all, which is
    # not the same as a zero: the cells stay empty rather than claiming a
    # break-even that was never observed.
    known = raw["invested"] is not None or raw["won"] is not None
    invested = _f(raw["invested"]) or 0.0
    won = _f(raw["won"]) or 0.0
    profit = won - invested if known else None
    # Divided by the entries that cost something: a family half made of
    # freerolls would otherwise show a buy-in half its real level.
    avg_buyin = invested / raw["paid"] if raw["paid"] else 0.0
    bounty_won = _f(raw["bounty_won"]) or 0.0
    bounty_invested = _f(raw["bounty_invested"]) or 0.0
    prize_won = _f(raw["prize_won"]) or 0.0
    prize_invested = _f(raw["prize_invested"]) or 0.0
    depths = _exit_depths(exits, raw.get("family"))

    return {
        "family": raw.get("family"),
        "tournaments": n,
        "reentries": raw["reentries"] or 0,
        "avg_buyin": avg_buyin or None,
        "avg_field": _div(raw["field_sum"], raw["field_n"]),
        "hands": raw["hands"],
        "played": _duration(raw["played_seconds"]),
        "invested": invested if known else None,
        "won": won if known else None,
        "profit": profit,
        "profit_buyins": profit / avg_buyin if avg_buyin and profit is not None else None,
        "roi": profit / invested if invested and profit is not None else None,
        "rake": _f(raw["rake"]),
        "itm_n": raw["itm"],
        "itm_pct": _div(raw["itm"], n),
        "deep_pct": _div(raw["deep_runs"], n),
        "final_tables": raw["final_tables"],
        "median_finish": _f(raw["median_finish_pct"]),
        "median_exit": median(depths) if depths else None,
        "bounty_share_buyin": _div(bounty_invested, invested),
        "bounty_won": bounty_won,
        "bounty_balance": bounty_won - bounty_invested,
        "prize_balance": prize_won - prize_invested,
        "bounty_share_won": _div(bounty_won, won),
        "vpip": _div(raw["vpip_hands"], raw["hands"]),
        "pfr": _div(raw["pfr_hands"], raw["hands"]),
        "flop": _div(raw["flop_hands"], raw["hands"]),
        "three_bet": _div(raw["three_bet"], raw["three_bet_opp"]),
    }


_SUMMABLE = (
    "tournaments", "paid", "reentries", "itm", "cashed", "bountied", "final_tables",
    "deep_runs", "field_sum", "field_n", "invested", "prize_invested",
    "bounty_invested", "rake", "won", "prize_won", "bounty_won",
    "played_seconds", "hands", "vpip_hands", "pfr_hands", "flop_hands",
    "three_bet", "three_bet_opp",
)


def _formats_total(raws: list[dict], meta: dict) -> dict:
    """
    The counters of every format added together.

    Only counters are summed. The median finish comes from the session
    aggregate, because a median of medians is not a median of anything.
    """
    total = {key: sum(_f(r[key]) or 0 for r in raws) for key in _SUMMABLE}
    total["family"] = None
    total["median_finish_pct"] = meta["median_finish_pct"]
    return total


def _sheet_formats(wb: Workbook, data: dict) -> dict:
    """Formats side by side, never added up. Returns the chart's anchors."""
    raws = data["formats"]
    views = [_format_view(raw, data["exits"]) for raw in raws]
    total = _format_view(_formats_total(raws, data["meta"]), data["exits"])

    has_bounty = any(r["family"] in BOUNTY_FAMILIES for r in raws)
    columns = len(raws) + 2                    # label + formats + toutes formes
    note_col = columns + 1

    ws = wb.create_sheet("Formats")
    ws.sheet_properties.tabColor = TAB_KEY
    _printable(ws)
    _autosize(ws, {"A": 30, **{get_column_letter(i): 15 for i in range(2, columns + 1)},
                   get_column_letter(note_col): 52})

    ws["A1"] = "Vanilla, knockout et mystery, séparés"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Trois structures de gains différentes : les additionner produit un ROI qui ne "
        "correspond à aucun jeu réel. Le mystery est le plus trompeur — sa prime n'est "
        "tirée qu'à partir d'un certain stade, donc son rendement vit presque entièrement "
        "dans des sessions que tu ne joues pas souvent."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=note_col)
    ws.row_dimensions[2].height = 30

    header = 4
    _header_row(ws, header, ["", *[FORMAT_LABELS.get(r["family"], r["family"]) for r in raws],
                             "Toutes formes", ""])
    for col, raw in enumerate(raws, start=2):
        _note(ws, header + 1, FORMAT_NOTES.get(raw["family"], ""), column=col)
        ws.cell(row=header + 1, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[header + 1].height = 26

    row = header + 2
    profit_row = row
    hidden = False                  # a whole section skipped, not just its banner
    for label, key, fmt, note in _FORMAT_ROWS:
        if label == "§":
            hidden = key == "PRIMES" and not has_bounty
            if not hidden:
                _banner(ws, row, key, columns)
                row += 1
            continue
        if hidden:
            continue

        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        for col, view in enumerate([*views, total], start=2):
            # A bounty line on a vanilla column is empty, not zero: there is no
            # bounty pool to have a balance in.
            blank = key.startswith(("bounty", "prize_balance")) and view["family"] == "vanilla"
            cell = ws.cell(row=row, column=col, value=None if blank else view[key])
            cell.border = BOX
            cell.font = TOTAL_FONT if col == columns else VALUE_FONT
            if col == columns:
                cell.fill = TOTAL_FILL
            if fmt:
                cell.number_format = fmt
        if note:
            _note(ws, row, note, column=note_col)
        if key == "profit":
            profit_row = row
        row += 1

    ws.freeze_panes = f"B{header + 2}"
    return {"header": header, "profit_row": profit_row, "families": len(raws)}


# ── Comparatif ────────────────────────────────────────────────────────────────

# (label, key, sample key, format, signed format, note)
_COMPARED_STATS = [
    ("VPIP", "vpip", "hands", PCT, PCT_SIGNED, "mains jouées volontairement"),
    ("PFR", "pfr", "hands", PCT, PCT_SIGNED, "mains ouvertes par une relance"),
    ("Flop vu", "flop", "hands", PCT, PCT_SIGNED, ""),
    ("3-bet", "three_bet", "n_3bet_opp", PCT, PCT_SIGNED, "relances face à une ouverture"),
    ("Fold to 3-bet", "fold_to_3bet", "n_faced_3bet", PCT, PCT_SIGNED, "abandons après avoir ouvert"),
    ("C-bet flop", "cbet_flop", "n_cbet_opp", PCT, PCT_SIGNED, "continuations en tant qu'agresseur"),
    ("WTSD", "wtsd", "n_flop", PCT, PCT_SIGNED, "abattages après avoir vu le flop"),
    ("W$SD", "wsd", "n_showdown", PCT, PCT_SIGNED, "abattages gagnés"),
    ("Facteur d'agression", "aggression_factor", "n_calls", RATIO, RATIO_SIGNED,
     "(mises + relances) / suivis"),
    ("Fréquence d'agression", "aggression_freq", "n_postflop", PCT, PCT_SIGNED, "postflop"),
]


def _sheet_comparatif(wb: Workbook, data: dict) -> None:
    session, lifetime = data["rates"], data["rates_lifetime"]

    ws = wb.create_sheet("Comparatif")
    ws.sheet_properties.tabColor = TAB_KEY
    _autosize(ws, {"A": 24, "B": 14, "C": 14, "D": 12, "E": 22, "F": 40})

    ws["A1"] = "Session vs référence"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"{session['hands']} mains de MTT sur la session, comparées à tes "
        f"{lifetime['hands']} mains de MTT jouées jusqu'à cette date. La colonne « base » donne "
        "le nombre d'occasions derrière le chiffre de la session : c'est elle qui décide "
        "si un écart vaut la peine d'être regardé."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 30

    _header_row(ws, 4, ["Statistique", "Session", "Référence", "Écart", "Base", "Définition"])
    row = 5
    for label, key, sample_key, fmt, signed, note in _COMPARED_STATS:
        here, ref = _f(session[key]), _f(lifetime[key])
        delta = here - ref if here is not None and ref is not None else None
        sample = session[sample_key]
        _data_row(ws, row, [label, here, ref, delta, _reliability(sample), note],
                  [None, fmt, fmt, signed, None, None])
        for col in (5, 6):
            ws.cell(row=row, column=col).font = NOTE_FONT
            ws.cell(row=row, column=col).border = Border()
        if not sample or sample < SAMPLE_INDICATIVE:
            for col in (2, 4):
                ws.cell(row=row, column=col).font = NOTE_FONT
        row += 1

    ws.freeze_panes = "A5"
    _note(ws, row + 1,
          "Les lignes grisées reposent sur trop peu d'occasions pour être lues comme un écart "
          "de jeu. Elles ne sont pas fausses, elles sont vides d'information.")


# ── Profondeur ────────────────────────────────────────────────────────────────

_DEPTH_COLUMNS = [
    ("Profondeur", 18, None),
    ("Mains", 9, COUNT),
    ("% des mains", 12, PCT),
    ("Tapis moyen", 12, BB_UNIT),
    ("VPIP", 9, PCT),
    ("réf.", 9, PCT),
    ("PFR", 9, PCT),
    ("réf.", 9, PCT),
    ("3-bet", 9, PCT),
    ("Flop vu", 10, PCT),
    ("WTSD", 9, PCT),
    ("Net (bb)", 11, BB),
    ("Net / 100 mains", 15, BB),
]


def _sheet_profondeur(wb: Workbook, data: dict) -> int:
    """
    The same rates, split by how deep the stack was.

    A tournament is three games in a row, and a single VPIP averages them into
    a number that describes none of them. Split by depth, the session becomes
    comparable to the reference line by line — and the row that is off is the
    one to work on.
    """
    rows = data["depths"]
    reference = {r["ord"]: r for r in data["depths_lifetime"]}
    total_hands = sum(r["hands"] for r in rows) or 1

    ws = wb.create_sheet("Profondeur")
    ws.sheet_properties.tabColor = TAB_KEY
    _autosize(ws, {get_column_letter(i): w for i, (_, w, _) in enumerate(_DEPTH_COLUMNS, 1)})

    ws["A1"] = "Jeu par profondeur de tapis"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Le tapis est mesuré en début de main, en big blinds. « réf. » est ton historique "
        "de MTT jusqu'à cette session, sur la même tranche de profondeur. Expresso et "
        "assimilés sont exclus : leur structure ne traverse pas ces tranches."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:M2")

    _header_row(ws, 4, [label for label, _, _ in _DEPTH_COLUMNS])
    formats = [fmt for _, _, fmt in _DEPTH_COLUMNS]

    row = 4
    for row, d in enumerate(rows, start=5):
        ref = reference.get(d["ord"], {})
        net_bb = _f(d["net_bb"])
        _data_row(ws, row, [
            d["bucket"],
            d["hands"],
            d["hands"] / total_hands,
            _f(d["avg_depth"]),
            _f(d["vpip"]), _f(ref.get("vpip")),
            _f(d["pfr"]), _f(ref.get("pfr")),
            _f(d["three_bet"]),
            _f(d["flop"]),
            _f(d["wtsd"]),
            net_bb,
            net_bb / d["hands"] * 100 if net_bb is not None and d["hands"] else None,
        ], formats)
        for col in (6, 8):
            ws.cell(row=row, column=col).font = NOTE_FONT

    ws.freeze_panes = "A5"
    if rows:
        ws.conditional_formatting.add(f"L5:L{row}", _profit_scale())
    _note(ws, row + 2,
          "Le net en bb d'une seule session est dominé par deux ou trois mains : lis les "
          "fréquences, pas le résultat. La ligne « moins de 10 bb » est celle où une erreur "
          "coûte le plus cher, parce qu'elle se joue sans postflop pour la rattraper.")
    return len(rows)


# ── Sorties ───────────────────────────────────────────────────────────────────

_EXIT_COLUMNS = [
    ("Tournoi", 34, None),
    ("Format", 11, None),
    ("Place", 8, COUNT),
    ("Champ", 9, COUNT),
    ("% du champ", 11, PCT),
    ("Niveau", 8, COUNT),
    ("Tapis (bb)", 11, BB_UNIT),
    ("Position", 10, None),
    ("Cartes", 10, None),
    ("Main (bb)", 11, BB),
    ("Abattage", 10, None),
    ("Gains", 11, MONEY),
]


def _sheet_sorties(wb: Workbook, data: dict) -> None:
    """
    One row per tournament: the last hand played, and the stack carried into it.

    Sorted deepest first, because that is the order of interest. Busting at
    7 bb is the format working as intended; busting at 35 bb means a pot was
    entered, and that pot is a decision worth reopening. This is the sheet that
    survives the variance — the stack you had is a fact, the result is a draw.
    """
    # Multi-table tournaments only: an Expresso ends when its structure says so,
    # so listing one here would put a decision nobody made at the top of a sheet
    # about decisions.
    exits = [e for e in data["exits"] if e["family"] in sessions.MTT_FAMILIES]
    depths = _exit_depths(data["exits"])

    ws = wb.create_sheet("Sorties")
    ws.sheet_properties.tabColor = TAB_KEY
    _autosize(ws, {get_column_letter(i): w for i, (_, w, _) in enumerate(_EXIT_COLUMNS, 1)})

    ws["A1"] = "Comment chaque MTT s'est terminé"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Sortie médiane à {median(depths):.1f} bb. "
        if depths else ""
    ) + (
        f"Les lignes en rouge sont les sorties au-dessus de {sessions.DEEP_EXIT_BB} bb : "
        "le tapis permettait encore de jouer, donc la main a été choisie. "
        "Expresso et assimilés ne figurent pas ici."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:L2")

    _header_row(ws, 4, [label for label, _, _ in _EXIT_COLUMNS])
    formats = [fmt for _, _, fmt in _EXIT_COLUMNS]

    row = 4
    for row, e in enumerate(exits, start=5):
        _data_row(ws, row, [
            e["name"] or e["tournament_id"],
            FORMAT_LABELS.get(e["family"], e["family"]),
            e["finish_position"],
            e["players_registered"],
            _f(e["finish_pct"]),
            e["level"],
            # A tournament won has a last hand but no exit depth: the stack at
            # that point is the whole table's chips, which measures nothing.
            None if e["won_tournament"] else _f(e["exit_bb"]),
            e["hero_position"],
            _cards(e["hero_cards"]),
            _f(e["net_bb"]),
            "oui" if e["went_to_showdown"] else "",
            _f(e["total_won"]),
        ], formats)
        if e["won_tournament"]:
            for col in range(1, len(_EXIT_COLUMNS) + 1):
                ws.cell(row=row, column=col).fill = TOTAL_FILL

    ws.freeze_panes = "A5"
    if exits:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(_EXIT_COLUMNS))}{row}"
        ws.conditional_formatting.add(
            f"G5:G{row}",
            CellIsRule(operator="greaterThanOrEqual",
                       formula=[str(sessions.DEEP_EXIT_BB)],
                       fill=WARN_FILL, font=WARN_FONT),
        )
    _note(ws, row + 2,
          "Le tapis est celui du début de la dernière main jouée : ce que tu avais en main "
          "au moment de la décision. Une ligne grisée est un tournoi gagné — il a bien une "
          "dernière main, mais ce n'est pas une sortie.")


# ── Tournois ──────────────────────────────────────────────────────────────────

_TOURNAMENT_COLUMNS = [
    ("Tournoi", 34, None),
    ("Format", 11, None),
    ("Vitesse", 11, None),
    ("Début", 16, TIME),
    ("Entrées", 9, COUNT),
    ("Buy-in", 11, MONEY),
    ("Joueurs", 9, COUNT),
    ("Place", 8, COUNT),
    ("% du champ", 11, PCT),
    ("Prize", 11, MONEY),
    ("Primes", 11, MONEY),
    ("Gains", 11, MONEY),
    ("Profit", 11, MONEY),
    ("Profit cumulé", 13, MONEY),
    ("ROI", 10, PCT_SIGNED),
    ("Durée", 9, None),
    ("Mains", 8, COUNT),
    ("Sortie (bb)", 11, BB_UNIT),
    ("VPIP", 8, PCT),
    ("PFR", 8, PCT),
    ("Flop", 8, PCT),
]

_TOURNAMENT_LABELS = [label for label, _, _ in _TOURNAMENT_COLUMNS]


def _sheet_tournois(wb: Workbook, data: dict) -> int:
    """Returns the number of tournament rows, for the chart to reference."""
    ws = wb.create_sheet("Tournois")
    ws.sheet_properties.tabColor = TAB_DETAIL
    _autosize(ws, {get_column_letter(i): w for i, (_, w, _) in enumerate(_TOURNAMENT_COLUMNS, 1)})

    _header_row(ws, 1, _TOURNAMENT_LABELS)
    formats = [fmt for _, _, fmt in _TOURNAMENT_COLUMNS]

    cumulative = 0.0
    row = 1
    for row, t in enumerate(data["tournaments"], start=2):
        buyin = _f(t["buyin_total"]) or 0.0
        won = _f(t["total_won"]) or 0.0
        profit = won - buyin
        cumulative += profit
        _data_row(ws, row, [
            t["name"] or t["tournament_id"],
            FORMAT_LABELS.get(t["family"], t["family"]),
            t["speed"],
            _local(t["started_at"]),
            t["entries"],
            buyin,
            t["players_registered"],
            t["finish_position"],
            _div(t["finish_position"], t["players_registered"]),
            _f(t["prize_won"]),
            _f(t["bounty_won"]),
            won,
            profit,
            cumulative,
            profit / buyin if buyin else None,
            _duration(t["duration_seconds"]),
            t["hands"],
            None if t["finish_position"] == 1 else _f(t["exit_bb"]),
            _f(t["vpip"]),
            _f(t["pfr"]),
            _f(t["flop"]),
        ], formats)

    total = row + 1
    totals = {
        "Tournoi": f"TOTAL — {len(data['tournaments'])} tournois",
        "Buy-in": _f(data["meta"]["total_invested"]) or 0.0,
        "Prize": _f(data["meta"]["prize_won"]),
        "Primes": _f(data["meta"]["bounty_won"]),
        "Gains": _f(data["meta"]["total_won"]) or 0.0,
        "Profit": _f(data["meta"]["net_profit"]) or 0.0,
        "Mains": data["meta"]["hands_count"],
    }
    _data_row(ws, total, [totals.get(label) for label in _TOURNAMENT_LABELS], formats)
    for col in range(1, len(_TOURNAMENT_COLUMNS) + 1):
        ws.cell(row=total, column=col).font = TOTAL_FONT
        ws.cell(row=total, column=col).fill = TOTAL_FILL

    ws.freeze_panes = "B2"
    if data["tournaments"]:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_TOURNAMENT_COLUMNS))}{row}"
        profit_col = get_column_letter(_TOURNAMENT_LABELS.index("Profit") + 1)
        ws.conditional_formatting.add(f"{profit_col}2:{profit_col}{row}", _profit_scale())
    return len(data["tournaments"])


# ── Positions ─────────────────────────────────────────────────────────────────

def _sheet_positions(wb: Workbook, data: dict) -> int:
    ws = wb.create_sheet("Positions")
    ws.sheet_properties.tabColor = TAB_DETAIL
    _autosize(ws, {"A": 12, "B": 10, "C": 12, "D": 10, "E": 10, "F": 10, "G": 14, "H": 12})

    _header_row(ws, 1, [
        "Position", "Mains", "% des mains", "VPIP", "PFR", "Flop vu",
        "Net (jetons)", "Net (BB)",
    ])
    formats = [None, COUNT, PCT, PCT, PCT, PCT, CHIPS, BB]

    total_hands = sum(p["hands"] for p in data["positions"]) or 1
    row = 1
    for row, p in enumerate(data["positions"], start=2):
        _data_row(ws, row, [
            p["position"], p["hands"], p["hands"] / total_hands,
            _f(p["vpip"]), _f(p["pfr"]), _f(p["flop"]),
            p["net_chips"], _f(p["net_bb"]),
        ], formats)

    ws.freeze_panes = "A2"
    _note(ws, row + 2, (
        "Le net en BB rapporte les jetons au niveau de blinds de chaque main : c'est la seule "
        "unité comparable entre le début et la fin d'un tournoi. Sur une seule session il reste "
        "dominé par quelques mains — les fréquences se lisent, le net se regarde."
    ))
    return len(data["positions"])


# ── Mains marquantes ──────────────────────────────────────────────────────────

def _sheet_mains(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Mains marquantes")
    ws.sheet_properties.tabColor = TAB_DETAIL
    _autosize(ws, {
        "A": 16, "B": 14, "C": 8, "D": 10, "E": 10, "F": 9,
        "G": 11, "H": 14, "I": 11, "J": 8, "K": 10, "L": 9,
    })
    labels = [
        "Heure", "Tournoi", "Niveau", "Position", "Cartes", "BB",
        "Tapis (bb)", "Net (jetons)", "Net (BB)", "Flop", "Abattage", "Gagnée",
    ]
    formats = [TIME, None, COUNT, None, None, CHIPS, BB_UNIT, CHIPS, BB, None, None, None]

    row = 1
    for title, hands in (
        ("Les 10 plus gros gains", data["best_hands"]),
        ("Les 10 plus grosses pertes", data["worst_hands"]),
    ):
        _banner(ws, row, title, len(labels))
        row += 1

        _header_row(ws, row, labels)
        row += 1
        for h in hands:
            _data_row(ws, row, [
                _local(h["hand_datetime"]),
                h["tournament_id"],
                h["level"],
                h["hero_position"],
                _cards(h["hero_cards"]),
                h["big_blind"],
                _f(h["depth_bb"]),
                h["net_chips"],
                _f(h["net_bb"]),
                "oui" if h["saw_flop"] else "",
                "oui" if h["went_to_showdown"] else "",
                "oui" if h["won_hand"] else "",
            ], formats)
            row += 1
        row += 2

    _note(ws, row, (
        "La colonne « tapis » donne la profondeur au début de la main : c'est elle qui dit si "
        "le pot était évitable."
    ))


# ── Graphiques ────────────────────────────────────────────────────────────────

def _sheet_graphiques(wb: Workbook, anchors: dict) -> None:
    ws = wb.create_sheet("Graphiques")
    ws.sheet_properties.tabColor = TAB_FREE
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Graphiques de la session"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 4

    tournaments = anchors["tournaments"]
    if tournaments:
        source = wb["Tournois"]
        column = _TOURNAMENT_LABELS.index("Profit cumulé") + 1
        curve = LineChart()
        curve.title = "Profit cumulé au fil de la session"
        curve.y_axis.title = "€"
        curve.x_axis.title = "Tournois, dans l'ordre de sortie"
        curve.height, curve.width = 9, 22
        curve.add_data(
            Reference(source, min_col=column, min_row=1, max_row=1 + tournaments),
            titles_from_data=True,
        )
        curve.set_categories(Reference(source, min_col=1, min_row=2, max_row=1 + tournaments))
        ws.add_chart(curve, "B3")

    formats = anchors["formats"]
    if formats and formats["families"] > 1:
        source = wb["Formats"]
        last = 1 + formats["families"]         # the "toutes formes" column is left out
        bars = BarChart()
        bars.type = "col"
        bars.title = "Profit par format"
        bars.y_axis.title = "€"
        bars.height, bars.width = 9, 10
        bars.add_data(
            Reference(source, min_col=1, max_col=last,
                      min_row=formats["profit_row"], max_row=formats["profit_row"]),
            from_rows=True, titles_from_data=True,
        )
        bars.set_categories(
            Reference(source, min_col=2, max_col=last,
                      min_row=formats["header"], max_row=formats["header"])
        )
        ws.add_chart(bars, "B22")

    depths = anchors["depths"]
    if depths:
        source = wb["Profondeur"]
        bars = BarChart()
        bars.type = "col"
        bars.title = "Mains par profondeur de tapis"
        bars.y_axis.title = "Mains"
        bars.height, bars.width = 9, 10
        bars.add_data(Reference(source, min_col=2, min_row=4, max_row=4 + depths),
                      titles_from_data=True)
        bars.set_categories(Reference(source, min_col=1, min_row=5, max_row=4 + depths))
        ws.add_chart(bars, "L22")

    positions = anchors["positions"]
    if positions:
        source = wb["Positions"]
        bars = BarChart()
        bars.type = "col"
        bars.title = "Résultat net par position"
        bars.y_axis.title = "Big blinds"
        bars.height, bars.width = 9, 22
        bars.add_data(Reference(source, min_col=8, min_row=1, max_row=1 + positions),
                      titles_from_data=True)
        bars.set_categories(Reference(source, min_col=1, min_row=2, max_row=1 + positions))
        ws.add_chart(bars, "B41")


# ── Notes ─────────────────────────────────────────────────────────────────────

_NOTE_PROMPTS = [
    ("Sorties à revoir (voir l'onglet Sorties, du haut vers le bas)", 5),
    ("Ce qui a bien fonctionné", 4),
    ("Erreurs identifiées", 5),
    ("État mental / tilt", 4),
    ("Adversaires et lectures", 4),
    ("À travailler d'ici la prochaine session", 5),
]


def _sheet_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Notes")
    ws.sheet_properties.tabColor = TAB_FREE
    ws.sheet_view.showGridLines = False
    _autosize(ws, {"A": 4, "B": 110})

    ws["B1"] = "Notes de session"
    ws["B1"].font = TITLE_FONT
    ws["B2"] = "Onglet libre — c'est la partie que le tracker ne peut pas remplir à ta place."
    ws["B2"].font = SUBTITLE_FONT

    row = 4
    for prompt, lines in _NOTE_PROMPTS:
        cell = ws.cell(row=row, column=2, value=prompt)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left", indent=1)
        row += 1
        for _ in range(lines):
            blank = ws.cell(row=row, column=2, value=None)
            blank.border = BOX
            blank.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1


# ── Workbook ──────────────────────────────────────────────────────────────────

def build_workbook(data: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)            # the default empty sheet

    _sheet_resume(wb, data)
    formats = _sheet_formats(wb, data) if data["formats"] else None
    _sheet_comparatif(wb, data)
    depths = _sheet_profondeur(wb, data)
    _sheet_sorties(wb, data)
    tournaments = _sheet_tournois(wb, data)
    positions = _sheet_positions(wb, data)
    _sheet_mains(wb, data)
    _sheet_graphiques(wb, {
        "tournaments": tournaments,
        "formats": formats,
        "depths": depths,
        "positions": positions,
    })
    _sheet_notes(wb)

    wb.active = 0
    return wb


def _filename(meta: dict) -> str:
    profit = _f(meta["net_profit"]) or 0.0
    return _safe(f"{meta['session_id']}_{profit:+.2f}EUR") + ".xlsx"


def _drop_previous(conn, session_id: str, keep: str, out_dir: str) -> None:
    """
    Remove the workbook a previous export left behind under a different name.

    The profit is part of the filename, so a session re-exported after a late
    summary lands under a new one. Deleting is restricted to the path this
    table recorded, inside the export directory, with an .xlsx extension.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT filepath FROM session_exports WHERE session_id = %s", (session_id,))
        row = cur.fetchone()

    if not row or not row[0] or row[0] == keep:
        return

    previous = row[0]
    inside = os.path.commonpath([os.path.abspath(previous), os.path.abspath(out_dir)])
    if inside != os.path.abspath(out_dir) or not previous.endswith(".xlsx"):
        return
    if os.path.exists(previous):
        os.remove(previous)
        print(f"[Export] remplace {os.path.basename(previous)}")


def _update_ledger(path: str, meta: dict) -> bool:
    """
    Write the session into the bankroll spreadsheet.

    False means "busy, come back later" and nothing else: the caller then
    leaves the session unrecorded so the next pass tries again. Any other
    failure is reported and swallowed — a spreadsheet that cannot be filled is
    a reason to look at it, never a reason to hold up the export.
    """
    name = os.path.basename(path)
    try:
        row = bankroll.record_session(
            path,
            session_id=meta["session_id"],
            day=_local(meta["session_start"]).date(),
            tournaments=meta["tournaments_count"],
            seconds=(meta["session_end"] - meta["session_start"]).total_seconds(),
            invested=_f(meta["total_invested"]) or 0.0,
            won=_f(meta["total_won"]) or 0.0,
        )
        print(f"[Bankroll] {name} — ligne {row}")
        return True
    except bankroll.LedgerLocked:
        print(f"[Bankroll] {name} ouvert dans LibreOffice — session reprise au prochain passage")
        return False
    except FileNotFoundError:
        # Presque toujours la même cause dans un conteneur : le fichier existe
        # sur le NAS, mais hors du seul dossier monté. Le dire ici épargne une
        # heure de recherche à côté de la plaque.
        print(f"[Bankroll] introuvable : {path}")
        print("[Bankroll] le conteneur ne voit que le dossier d'export — placer le fichier "
              "dedans, ou monter son dossier dans docker-compose.yml")
        return True
    except (bankroll.LedgerError, OSError) as exc:
        print(f"[Bankroll] {name} non mis à jour : {exc}")
        return True


def export_session(conn, session_id: str, out_dir: str, *, gap_minutes: int,
                   ledger: str | None = None, ledger_player: str | None = None,
                   ledger_write: bool = True) -> str | None:
    data = sessions.fetch_session(conn, session_id, gap_minutes=gap_minutes)
    if data is None:
        print(f"[Export] session inconnue : {session_id}")
        return None

    meta = data["meta"]
    # The ledger follows one player: the watch directory holds the backups of
    # several machines, and a bankroll that mixed two accounts would describe
    # neither.
    book = ledger if ledger and ledger_player in (None, meta["player"]) else None
    # Read even when writing is off. Rebuilding history must not add rows to a
    # file the player keeps by hand, but there is no reason for the workbook it
    # produces to open on an empty bankroll cell.
    if book:
        data["bankroll_start"] = bankroll.opening_for(
            book, session_id, _local(meta["session_start"]).date()
        )

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, _filename(meta))

    # Written before the old one is removed: a save that fails must not leave
    # the session with no workbook at all.
    build_workbook(data).save(path)

    # Recorded only once the ledger has had it. Leaving the session out of
    # session_exports is what brings it back on the next tick, which is the
    # whole retry mechanism for a spreadsheet that was open at the wrong time.
    if book and ledger_write and not _update_ledger(book, meta):
        return path

    _drop_previous(conn, session_id, path, out_dir)
    sessions.record_export(conn, meta, path)
    conn.commit()

    print(f"[Export] {os.path.basename(path)}"
          f"  ({meta['tournaments_count']} tournois, {meta['hands_count']} mains)")
    return path


def export_due(conn, out_dir: str, *, gap_minutes: int, lookback_days: int, grace_hours: int,
               ledger: str | None = None, ledger_player: str | None = None,
               ledger_write: bool = True) -> int:
    """Every session that is over and not yet on disk. Used by the watcher tick."""
    due = sessions.due_sessions(
        conn, gap_minutes=gap_minutes, lookback_days=lookback_days, grace_hours=grace_hours
    )
    for meta in due:
        export_session(conn, meta["session_id"], out_dir, gap_minutes=gap_minutes,
                       ledger=ledger, ledger_player=ledger_player,
                       ledger_write=ledger_write)
    return len(due)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    what = parser.add_mutually_exclusive_group()
    what.add_argument("--last", action="store_true", help="la session la plus récente")
    what.add_argument("--session", metavar="ID", help="une session précise")
    what.add_argument("--rebuild-all", action="store_true", help="tout l'historique")
    what.add_argument("--since", metavar="AAAA-MM-JJ", help="les sessions depuis cette date")
    what.add_argument("--list", action="store_true", help="lister les sessions sans rien écrire")
    parser.add_argument("--out", default=os.environ.get("EXPORT_DIR"), help="dossier de destination")
    parser.add_argument("--gap", type=int,
                        default=int(os.environ.get("SESSION_GAP_MINUTES", sessions.DEFAULT_GAP_MINUTES)),
                        help="minutes d'inactivité qui séparent deux sessions")
    parser.add_argument("--db", default=os.environ.get("DB_URL"), help="DSN PostgreSQL")
    parser.add_argument("--bankroll", metavar="FICHIER.ods",
                        default=os.environ.get("BANKROLL_FILE"),
                        help="fichier de bankroll à tenir à jour")
    parser.add_argument("--bankroll-player", metavar="PSEUDO",
                        default=os.environ.get("BANKROLL_PLAYER"),
                        help="ne porter au fichier que les sessions de ce joueur")
    parser.add_argument("--no-bankroll", action="store_true",
                        help="ne pas toucher au fichier de bankroll")
    parser.add_argument("--bankroll-history", action="store_true",
                        help="porter aussi les sessions d'une reprise d'historique "
                             "(--since, --rebuild-all) dans le fichier de bankroll")
    args = parser.parse_args(argv)
    # Résolu ici pour le CLI comme pour le watcher, par la même fonction.
    ledger = None if args.no_bankroll else bankroll.default_path(args.out, args.bankroll)
    ledger_write = True

    if not args.db:
        parser.error("DB_URL manquant (variable d'environnement ou --db)")
    if not args.out and not args.list:
        parser.error("EXPORT_DIR manquant (variable d'environnement ou --out)")

    conn = psycopg2.connect(args.db)
    try:
        if args.list:
            for meta in sessions.list_sessions(conn, gap_minutes=args.gap):
                start = _local(meta["session_start"])
                print(f"{meta['session_id']:<32} {start:%d/%m/%Y %H:%M}  "
                      f"{meta['tournaments_count']:>3} tournois  "
                      f"{_f(meta['net_profit']) or 0:+8.2f} €")
            return 0

        if args.session:
            targets = [args.session]
        elif args.last:
            latest = sessions.latest_session_id(conn, gap_minutes=args.gap)
            targets = [latest] if latest else []
        elif args.rebuild_all or args.since:
            since = datetime.strptime(args.since, "%Y-%m-%d") if args.since else None
            targets = [m["session_id"] for m in sessions.list_sessions(conn, gap_minutes=args.gap, since=since)]
            # Rebuilding writes workbooks; it only writes rows into the ledger
            # when asked for. That file is the player's own, and pouring months
            # of past sessions into it has to be a decision, not a side effect
            # of regenerating a folder of .xlsx. Read either way, so each
            # workbook opens on the bankroll of its own evening.
            #
            # Sessions come out oldest first (list_sessions orders on
            # session_start), which is what the bankroll column needs: it
            # chains on the row above it.
            ledger_write = args.bankroll_history
            if ledger and not ledger_write:
                print("[Bankroll] reprise d'historique : le fichier est lu, jamais modifié"
                      " — utiliser --bankroll-history pour y porter ces sessions")
            elif ledger:
                print("[Bankroll] reprise d'historique portée au fichier, la plus ancienne d'abord")
        else:
            count = export_due(
                conn, args.out,
                gap_minutes=args.gap,
                lookback_days=int(os.environ.get("EXPORT_LOOKBACK_DAYS", 7)),
                grace_hours=int(os.environ.get("EXPORT_GRACE_HOURS", 6)),
                ledger=ledger, ledger_player=args.bankroll_player,
                ledger_write=ledger_write,
            )
            print(f"[Export] {count} session(s) traitée(s)")
            return 0

        if not targets:
            print("[Export] aucune session à traiter")
            return 0
        print(f"[Export] {len(targets)} session(s) → {os.path.abspath(args.out)}")
        for session_id in targets:
            export_session(conn, session_id, args.out, gap_minutes=args.gap,
                           ledger=ledger, ledger_player=args.bankroll_player,
                           ledger_write=ledger_write)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
